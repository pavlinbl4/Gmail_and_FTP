"""
скрипт собирающий информацию о всех моих фотографиях в агенстве ТАСС
и сохраняющий эту информацию на отдельной вкладке xlsx файла в соответствии с датой запуска
"""

import os
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime

headers = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:98.0) Gecko/20100101 Firefox/98.0",
    "Host": "www.tassphoto.com",
    "Referer": "https://www.tassphoto.com/ru/asset/fullTextSearch/search/%D0%A1%D0%B5%D0%BC%D0%B5%D0%BD+%D0%9B%D0%B8%D1%85%D0%BE%D0%B4%D0%B5%D0%B5%D0%B2/page/1",
    "accept": "*/*",
    "Cookie": "IUSR_GUEST_savedComboSearchParams=a%3A11%3A%7Bs%3A15%3A%22selected_topics%22%3Ba%3A0%3A%7B%7Ds%3A13%3A%22selected_date%22%3Bs%3A0%3A%22%22%3Bs%3A22%3A%22selected_dateTakenFrom%22%3Bs%3A0%3A%22%22%3Bs%3A20%3A%22selected_dateTakenTo%22%3Bs%3A0%3A%22%22%3Bs%3A16%3A%22selected_regions%22%3Ba%3A0%3A%7B%7Ds%3A15%3A%22selected_stocks%22%3Ba%3A0%3A%7B%7Ds%3A21%3A%22selected_photographer%22%3Bs%3A0%3A%22%22%3Bs%3A6%3A%22search%22%3Bs%3A31%3A%22%D0%B5%D0%B2%D0%B3%D0%B5%D0%BD%D0%B8%D0%B9+%D0%BF%D0%B0%D0%B2%D0%BB%D0%B5%D0%BD%D0%BA%D0%BE%22%3Bs%3A20%3A%22selected_orientation%22%3Bs%3A3%3A%22all%22%3Bs%3A17%3A%22selected_doctypes%22%3Ba%3A0%3A%7B%7Ds%3A12%3A%22selected_ref%22%3Bs%3A0%3A%22%22%3B%7D; IUSR_GUEST_savedSearchParams=a%3A1%3A%7Bs%3A6%3A%22search%22%3Bs%3A31%3A%22%D0%B5%D0%B2%D0%B3%D0%B5%D0%BD%D0%B8%D0%B9+%D0%BF%D0%B0%D0%B2%D0%BB%D0%B5%D0%BD%D0%BA%D0%BE%22%3B%7D; PHPSESSID=bvaafujetdmd9bt0ds3s1htuk4; uiPrefs=a%3A4%3A%7Bs%3A21%3A%22isBottomLightboxShown%22%3Bb%3A0%3Bs%3A23%3A%22isLayoutLeftColumnShown%22%3Bb%3A1%3Bs%3A24%3A%22isLayoutRightColumnShown%22%3Bb%3A0%3Bs%3A50%3A%22pagination_asset_fullTextSearch_nb_thumbs_per_page%22%3Bs%3A2%3A%2220%22%3B%7D"
}


def get_html(link):
    req = requests.get(link, headers=headers)
    return req.text


def get_soup(html):
    soup = BeautifulSoup(html, 'lxml')
    return soup


def get_page_nambers(url):  # количество страниц с которых нужно  собрать информацию
    soup = get_soup(get_html(f'{url}1'))
    images_online = int(str(soup.select(".result-counter#nb-result"))[42:47])
    page_number = images_online // 20 + 1
    return page_number, images_online


def create_xlsx():
    today = datetime.now().strftime("%Y-%m-%d")
    if os.path.exists(f'{report_folder}/all_TASS_images.xlsx'):
        wb = load_workbook(f'{report_folder}/all_TASS_images.xlsx')  # файл есть и открываю его
        ws = wb.create_sheet(today)  # добавляю новую таблицу
    else:
        wb = Workbook()  # если файда еще нет
        ws = wb.active  # если файда еще нет
        ws.title = today  # если файда еще нет

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10  # задаю ширину колонки
    ws.column_dimensions['D'].width = 110
    ws.column_dimensions['E'].width = 50
    return ws, wb


def main():
    page_number, images_online = get_page_nambers(url)
    ws, wb = create_xlsx()
    count = 0

    for n in range(1, page_number + 1):  # количество страниц для анализа  - page_number + 1
        link = f'{url}{n}'
        soup = get_soup(get_html(link))
        thumbs_data = soup.find('ul', id="mosaic").find_all('div', class_="thumb-content thumb-width thumb-height")
        images_on_page = len(soup.find('ul', id="mosaic").find_all('a', class_="zoom"))
        for i in range(images_on_page):
            count += 1
            image_date = thumbs_data[i].find(class_="date").text
            image_id = thumbs_data[i].find(class_="title").text
            image_title = thumbs_data[i].find('p').text
            image_caption = soup.find('ul', id="mosaic").find_all(class_="thumb-text")[i].text.strip().split('\n')[
                -1].lstrip().replace(' Семен Лиходеев/ТАСС', '').replace(' Фото ИТАР-ТАСС/ Семен Лиходеев','')
            image_link = soup.find('ul', id="mosaic").find_all('a', class_="zoom")[i].find('img').get('src')
            print(count, image_id, image_date)
            print(image_title)
            print(image_caption)
            print(image_link)
            ws[f'A{count}'] = images_online + 1 - count
            ws[f'B{count}'] = image_id
            ws[f'C{count}'] = image_date
            ws[f'D{count}'] = image_caption
            ws[f'E{count}'] = image_link
    wb.save(f'{report_folder}/all_TASS_images.xlsx')


url = 'https://www.tassphoto.com/ru/asset/fullTextSearch/search/%D0%A1%D0%B5%D0%BC%D0%B5%D0%BD%20%D0%9B%D0%B8%D1%85%D0%BE%D0%B4%D0%B5%D0%B5%D0%B2/page/'

report_folder = "/Volumes/big4photo/Documents/TASS/Tass_data"

if __name__ == '__main__':
    main()
