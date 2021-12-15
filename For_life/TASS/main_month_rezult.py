"""
основной модуль для добавления информации в файл отчета

информация за месяц - сколько какое фото продано
добавляю вкладку в общий файл работает правильно
декабрь 2021 окончательный вариант
- добавить функцию удаления исходного файла или его переименования
данный скрипт работает с использованием work_with_XLXS и report_date_from_file
файлов

задача - присоединить модуль скачивания превью и создать модуль выбора пути
к файлу отчета
"""

import work_with_XLXS
import report_date_from_file
from openpyxl import load_workbook


main_report = "/Volumes/big4photo/Documents/TASS/Tass_total_report.xlsx" # файл куда сохранятеся вся информация
file_to_work = '/Volumes/big4photo/Downloads/Павленко Евгений.xlsx'  # файл из которого будет извлекаться информация
report_date = report_date_from_file.get_report_date(file_to_work) # получаю неазвание месяца из отчета
photos = work_with_XLXS.get_publications(file_to_work)

wb = load_workbook(filename=main_report, read_only=False)
ws_month_number = wb.create_sheet(report_date, 0)
ws_month_number.cell(row =1,column=1).value = "photo_id"
ws_month_number.cell(row =1,column=2).value = "income"
ws_month_number.cell(row =1,column=3).value = "sold times"

kkeys =[i for i in photos.keys()]
for i in range(len(photos)):
    ws_month_number.cell(row=2 + i, column=1).value = kkeys[i]
    ws_month_number.cell(row=2 + i, column=2).value = sum(photos[kkeys[i]]) # суммирую доход по снимкам
    ws_month_number.cell(row=2 + i, column=3).value = len(photos[kkeys[i]])



wb.save()
wb.close()
print(report_date)