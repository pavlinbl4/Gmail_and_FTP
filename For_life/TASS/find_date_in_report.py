"""
задача находить в таблице нужгые поля по слову, чтоб мой скрипт для таса мог работать со старыми отчетами
"""

from openpyxl import load_workbook
import re

file_name = '/Volumes/big4photo/Downloads/Павленко-фев.xlsx'


def get_report_date(file_name):  # получаю дату отчета в виде строки
    wb = load_workbook(file_name)
    sheet = wb.active
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value != None and str(cell.value).startswith('Период'):
                report_date = cell.value.split(': ')[1]
                print(report_date)


# get_report_date(file_name)

def gen_row(file_name):  # функция определяет номер строки с которой начинается ввод данных в табицу
    wb = load_workbook(file_name)
    sheet = wb.active
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'ID фото ':
                print(cell.column)
                return cell.column

gen_row(file_name)
