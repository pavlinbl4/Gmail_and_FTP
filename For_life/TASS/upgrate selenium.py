from selenium import webdriver
import time
import requests
import openpyxl
import os
from selenium.webdriver.common.by import By

options = webdriver.ChromeOptions()
options.add_argument(
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4200.0 Iron Safari/537.36")
browser = webdriver.Chrome(options=options)

file_way = '/Volumes/big4photo/Downloads/Павленко.xlsx'
wb = openpyxl.load_workbook(file_way)
sheet = wb.active
photos = {}
x = 7
# report_date = report_date_from_file.get_report_date(file_way)

os.makedirs(f"/Volumes/big4photo/Documents/TASS/images/upgrate_selenium", exist_ok=True)

try:
    browser.get('https://www.tassphoto.com/ru')
    time.sleep(1)

    photo_id = (sheet.cell(row=x, column=4)).value
    while photo_id != None:
        # search_input = browser.find_element_by_id("userrequest")
        search_input = browser.find_element(By.ID,"userrequest")
        search_input.clear()
        search_input.send_keys(photo_id)
        # browser.find_element_by_id("search-submit").click()
        browser.find_element(By.ID,"search-submit").click()
        # picture = browser.find_element_by_css_selector("#mosaic .zoom img").get_attribute("src")
        picture = browser.find_element(By.CSS_SELECTOR,"#mosaic .zoom img").get_attribute("src")
        print(picture)
        get_image = requests.get(picture)
        with open(f"/Volumes/big4photo/Documents/TASS/images/upgrate_selenium/{photo_id}.jpg", 'wb') as img_file:
            img_file.write(get_image.content)
        x += 1
        photo_id = (sheet.cell(row=x, column=4)).value

    browser.close()
    browser.quit()
except Exception as ex:
    print(ex)
    browser.close()
    browser.quit()