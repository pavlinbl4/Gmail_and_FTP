"""
скрипт собирающий информацию о всех моих фотографиях в агенстве ТАСС
и сохраняющий эту информацию на отдельной вкладке xlsx файла в соответствии с датой запуска
"""

import os
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup

options = webdriver.ChromeOptions()
options.add_argument(
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4200.0 Iron Safari/537.36")
options.add_argument("--disable-blink-features=AutomationControlled")
options.headless = True

browser = webdriver.Chrome(options=options)



def get_html(link):
    browser.get('https://www.tassphoto.com/ru')
    WebDriverWait(browser, 10).until(
        EC.presence_of_element_located((By.ID, "userrequest"))
    )
    search_input = browser.find_element(By.ID, "userrequest")
    search_input.clear()
    search_input.send_keys('Семен Лиходеев')
    search_input.send_keys(Keys.ENTER)
    browser.get(link)
    html = browser.page_source

    return html


def get_soup(html):
    soup = BeautifulSoup(html, 'lxml')
    return soup


def get_page_nambers(url):  # количество страниц с которых нужно  собрать информацию
    soup = get_soup(get_html(f'{url}1'))
    # print(soup)
    # print(soup.select(".result-counter#nb-result"))
    images_online = int(str(soup.select(".result-counter#nb-result"))[42:47])
    page_number = images_online // 20 + 1
    return page_number, images_online

def create_columns_names(ws):
    ws[f'A1'] = 'images_online'
    ws[f'B1'] = 'image_id'
    ws[f'C1'] = 'image_date'
    ws[f'D1'] = 'image_caption'
    ws[f'E1'] = 'image_link'

def create_xlsx():
    today = datetime.now().strftime("%Y-%m-%d")
    if os.path.exists(f'{report_folder}/all_TASS_images.xlsx'):
        wb = load_workbook(f'{report_folder}/all_TASS_images.xlsx')  # файл есть и открываю его
        ws = wb.create_sheet(today)  # добавляю новую таблицу
        create_columns_names(ws)
    else:
        wb = Workbook()  # если файда еще нет
        ws = wb.active  # если файда еще нет
        ws.title = today  # если файда еще нет
        create_columns_names(ws)

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10  # задаю ширину колонки
    ws.column_dimensions['D'].width = 110
    ws.column_dimensions['E'].width = 50
    return ws, wb


def main():
    page_number, images_online = get_page_nambers(url)
    ws, wb = create_xlsx()
    count = 1

    for n in range(1, page_number + 1):  # количество страниц для анализа  - page_number + 1
        link = f'{url}{n}'

        soup = get_soup(get_html(link))
        thumbs_data = soup.find('ul', id="mosaic").find_all('div', class_="thumb-content thumb-width thumb-height")
        images_on_page = len(soup.find('ul', id="mosaic").find_all('a', class_="zoom"))
        for i in range(images_on_page):
            count += 1
            image_date = thumbs_data[i].find(class_="date").text
            image_id = thumbs_data[i].find(class_="title").text
            image_title = thumbs_data[i].find('p').text
            image_caption = soup.find('ul', id="mosaic").find_all(class_="thumb-text")[i].text.strip().split('\n')[
                -1].lstrip().replace(' Семен Лиходеев/ТАСС', '').replace(' Фото ИТАР-ТАСС/ Семен Лиходеев','')
            image_link = soup.find('ul', id="mosaic").find_all('a', class_="zoom")[i].find('img').get('src')
            print(count - 1, image_id, image_date)
            print(image_title)
            print(image_caption)
            print(image_link)
            ws[f'A{count}'] = images_online + 1 - count
            ws[f'B{count}'] = image_id
            ws[f'C{count}'] = image_date
            ws[f'D{count}'] = image_caption
            ws[f'E{count}'] = image_link
    wb.save(f'{report_folder}/all_TASS_images.xlsx')
    wb.close()




url = 'https://www.tassphoto.com/ru/asset/fullTextSearch/search/%D0%A1%D0%B5%D0%BC%D0%B5%D0%BD%20%D0%9B%D0%B8%D1%85%D0%BE%D0%B4%D0%B5%D0%B5%D0%B2/page/'

report_folder = "/Volumes/big4photo/Documents/TASS/Tass_data"

if __name__ == '__main__':
    main()
    browser.close()
    browser.quit()
