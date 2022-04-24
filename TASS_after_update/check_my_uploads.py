"""проверяю, добавили ли мои фото на сайт, скрипт отрабатывает по crontab и
 заносит информацию в таблицу, если количество снимков изменилось, то запускаю скрипт
 all_images_new и потом добавляет свежие снимки в заданную папку """

from openpyxl import load_workbook
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import pandas as pd
import datacompy
import os
import requests
import re
import subprocess
import sys

options = webdriver.ChromeOptions()
options.headless = True  # фоновый режим
options.add_argument(
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4200.0 Iron Safari/537.36")
options.add_argument("--disable-blink-features=AutomationControlled")

browser = webdriver.Chrome(options=options)


def image_downloader(difference, last_date):
    folder = '/Volumes/big4photo/Documents/TASS/Tass_data/added_images'
    os.makedirs(f"{folder}/{last_date}", exist_ok=True)
    for i in range(len(difference)):
        image_url = difference.image_link.iloc[i]
        r = requests.get(image_url)
        image_name = re.findall(r'\d+(?=\.thw)', image_url)[0]
        print(image_name, image_url)
        with open(f"{folder}/{last_date}/{image_name}.jpg", 'wb') as download_file:
            for chunk in r.iter_content(9000):
                download_file.write(chunk)


def new_pictures_links(last_date, previous_date):
    pd.options.display.max_colwidth = 100
    photo_base = '/Volumes/big4photo/Documents/TASS/Tass_data/all_TASS_images.xlsx'
    last_df = pd.read_excel(photo_base, sheet_name=last_date)
    previos_df = pd.read_excel(photo_base, sheet_name=previous_date)
    difference = datacompy.Compare(previos_df, last_df,
                                   join_columns=['image_id', 'image_date', 'image_caption']).df2_unq_rows
    return difference


def get_html(link):
    browser.get('https://www.tassphoto.com/ru')
    WebDriverWait(browser, 10).until(
        EC.presence_of_element_located((By.ID, "userrequest"))
    )
    search_input = browser.find_element(By.ID, "userrequest")
    search_input.clear()
    search_input.send_keys('Семен Лиходеев')
    search_input.send_keys(Keys.ENTER)
    html = browser.page_source
    browser.close()
    browser.quit()
    return html


def parser():  # функция возвращает количество опубликованных на сайте снимков
    link = 'https://www.tassphoto.com/ru'
    html = get_html(link)
    soup = BeautifulSoup(html, 'lxml')
    return str(soup.select(".result-counter#nb-result"))[42:47]


def add_data(photos_count):
    file = "/Volumes/big4photo/Documents/TASS/Tass_data/TASS_photos.xlsx"
    book = load_workbook(file)
    ws = book.active
    last_row = ws.max_row
    old_value = ws.cell(row=last_row, column=2).value  # последние данные в таблице
    last_date = f'{ws.cell(row=last_row, column=1).value[:10]}' # дата внесения последних данных
    previous_date = f'{ws.cell(row=(last_row - 1), column=1).value[:10]}' # дата внесения предпоследних данных
    if ws.cell(row=last_row, column=2).value != photos_count:
        ws.cell(row=last_row + 1, column=2).value = photos_count
        ws.cell(row=last_row + 1, column=1).value = datetime.now().strftime('%Y-%m-%d %H:%M')
        new_images = int(photos_count) - int(old_value)  # количество добавленных фото
        ws.cell(row=last_row + 1, column=3).value = new_images
        print(f'добавлено {new_images} снимков')
        subprocess.run([sys.executable, 'all_images_new.py']) # запусакаю субпроцесс для записи даным по всем фотографиям в файл
        print('subprocess1 - ended')
        subprocess.run([sys.executable, 'download_fresh_images.py'])
        # difference = new_pictures_links(last_date, previous_date)
        # image_downloader(difference, last_date)

    else:
        print(f"no new images, now  {photos_count} - images online")

    book.save(file)
    book.close()


add_data(parser())
