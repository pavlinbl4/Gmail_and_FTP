import os
import requests
import json
import time
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from datetime import datetime, timedelta
from pathlib import Path

yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
folder = (Path.home() / "Documents" / "Kommersant" / "LentaRU").mkdir(parents=True, exist_ok=True)
report_folder = (Path.home() / "Documents" / "Kommersant" / "LentaRU")
print(yesterday)


def notification(message):
    title = "Готово"
    command = f'''
    osascript -e 'display notification "{message}" with title "{title}"'
    '''
    os.system(command)


def lenta_ru_time_converter(lenta_ru_time):
    return datetime.strptime(time.ctime(lenta_ru_time), '%a %b %d %H:%M:%S %Y').strftime('%Y-%m-%d')


def get_html(url):
    headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:97.0) Gecko/20100101 Firefox/97.0',
               'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
               'Connection': 'keep-alive'
               }
    r = requests.get(url, headers=headers)
    if r.status_code != 200:
        time.sleep(10)
        print("bad status code")
        r = requests.get(url, headers=headers)
    return r.text


def make_subfolder(photograf):
    folder = f"{report_folder}/{yesterday}/{photograf}"
    os.makedirs(folder, exist_ok=True)
    return folder


def download_image(folder, image_url):
    r = requests.get(image_url)
    image_name = image_url.split('/')[-1]
    with open(f"{folder}/{image_name}", 'wb') as download_file:
        for chunk in r.iter_content(9000):
            download_file.write(chunk)
    return f"{folder}/{image_name}"


def main():
    if os.path.exists(f'{report_folder}/Ъ_in_LentaRU.xlsx'):
        wb = load_workbook(f'{report_folder}/Ъ_in_LentaRU.xlsx')  # файл есть и открываю его
        ws = wb.create_sheet(yesterday)  # добавляю новую таблицу
    else:
        wb = Workbook()  # если файда еще нет
        ws = wb.active  # если файда еще нет
        ws.title = yesterday  # если файда еще нет

    ws.column_dimensions['C'].width = 50  # задаю ширину колонки
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 100

    url = 'https://lenta.ru/search/v2/process?from=0&size=50&sort=2&title_only=0&domain=1&modified,format=yyyy-MM-dd&query=фото Коммерсантъ'
    rezult = json.loads(get_html(url))
    count = 1
    for i in rezult['matches']:
        if lenta_ru_time_converter(i['pubdate']) == yesterday:
            count += 1
            photograf = i['text'].split('Фото:')[1].split('/')[0].strip()
            image_url = i['image_url']
            ws.row_dimensions[count].height = 100  # задаю высоту столбца
            print(i['title'])
            print(photograf, image_url, lenta_ru_time_converter(i['pubdate']))
            folder = make_subfolder(photograf)
            image_path = download_image(folder, image_url)

            img = Image(image_path)
            resize_height = img.height // 3  # уменьшая рарешение в два раза
            resize_width = img.width // 3  # уменьшая рарешение в два раза

            img.width = resize_width  # устанавливаю размер превью
            img.height = resize_height  # устанавливаю размер превью

            ws.add_image(img, f'B{count}')
            ws[f'A{count}'] = photograf

    wb.save(f'{report_folder}/Ъ_in_LentaRU.xlsx')
    notification("LentaRu completed")


if __name__ == '__main__':
    main()
